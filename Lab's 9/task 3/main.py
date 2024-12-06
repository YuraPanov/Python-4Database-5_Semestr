import openpyxl
from openpyxl.chart import PieChart, Reference


def update_salary_data(sheet):
    departments = [
        ("Бухгалтерия", 4),
        ("Отдел кадров", 9),
        ("Столовая", 11)
    ]

    for idx, (department, row) in enumerate(departments, start=2):
        sheet.cell(idx, 11).value = department
        sheet.cell(idx, 12).value = sheet.cell(row, 9).value


def create_pie_chart(sheet):
    pie = PieChart()

    labels = Reference(sheet, min_col=11, min_row=2, max_row=4)
    data = Reference(sheet, min_col=12, min_row=2, max_row=4)

    pie.add_data(data)
    pie.set_categories(labels)
    pie.title = "Зарплата по отделам"

    sheet.add_chart(pie, 'M1')


wb = openpyxl.load_workbook("table.xlsx")
sheet = wb.active

update_salary_data(sheet)

create_pie_chart(sheet)

wb.save("table_final.xlsx")

